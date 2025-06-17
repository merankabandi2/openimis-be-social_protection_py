import copy
import json
import logging
import uuid

import math
import pandas as pd
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db import transaction
from django.db import models
from django.db.models import Q, Value, Func, F
from django.db.models.functions import Concat
from pandas import DataFrame

from calculation.services import get_calculation_object
from core.services import BaseService
from core.signals import register_service_signal
from individual.models import IndividualDataSourceUpload, IndividualDataSource, Individual
from social_protection.apps import SocialProtectionConfig
from social_protection.models import (
    BenefitPlan,
    Beneficiary,
    BenefitPlanDataUploadRecords,
    GroupBeneficiary,
    BeneficiaryStatus,
    LocationBenefitPlanPaymentPoint,
    Project,
)

from social_protection.utils import load_dataframe, fetch_summary_of_valid_items, fetch_summary_of_broken_items
from social_protection.validation import (
    BeneficiaryValidation,
    BenefitPlanValidation,
    GroupBeneficiaryValidation,
    ProjectValidation,
)
from tasks_management.services import UpdateCheckerLogicServiceMixin, CheckerLogicServiceMixin, \
    crud_business_data_builder
from workflow.systems.base import WorkflowHandler
from workflow.util import result as WorkflowExecutionResult
from core.models import User
from core.services.utils import check_authentication as check_authentication, output_exception, \
    model_representation, output_result_success

from social_protection.apps import SocialProtectionConfig

logger = logging.getLogger(__name__)


class BenefitPlanService(BaseService, UpdateCheckerLogicServiceMixin):
    OBJECT_TYPE = BenefitPlan

    def __init__(self, user, validation_class=BenefitPlanValidation):
        super().__init__(user, validation_class)

    @register_service_signal('benefit_plan_service.create')
    def create(self, obj_data):
        return super().create(obj_data)

    @register_service_signal('benefit_plan_service.update')
    def update(self, obj_data):
        return super().update(obj_data)

    @register_service_signal('benefit_plan_service.delete')
    def delete(self, obj_data):
        obj_data = {k: v for k, v in obj_data.items() if k != 'user'}
        return super().delete(obj_data)

    @register_service_signal('benefit_plan_service.close')
    def close_benefit_plan(self, obj_data):
        from tasks_management.models import Task
        from tasks_management.apps import TasksManagementConfig
        from tasks_management.services import _get_std_task_data_payload, TaskService
        from social_protection.apps import SocialProtectionConfig
        benefit_plan = BenefitPlan.objects.filter(id=obj_data.get('id')).first()
        data = {'benefit_plan_id': benefit_plan.id}
        TaskService(self.user).create({
            'source': 'BenefitPlanService',
            'entity': benefit_plan,
            'status': Task.Status.RECEIVED,
            'executor_action_event': TasksManagementConfig.default_executor_event,
            'business_event': SocialProtectionConfig.benefit_plan_suspend,
            'data': _get_std_task_data_payload(data)
        })


class BeneficiaryService(BaseService, CheckerLogicServiceMixin):
    OBJECT_TYPE = Beneficiary

    def __init__(self, user, validation_class=BeneficiaryValidation):
        super().__init__(user, validation_class)

    def would_exceed_max_active_beneficiaries(self, benefit_plan_id, status, id=None):
        benefit_plan = BenefitPlan.objects.get(id=benefit_plan_id)
        if benefit_plan and status == "ACTIVE":
            max_active_beneficiaries = benefit_plan.max_beneficiaries
            active_beneficiaries = Beneficiary.objects.filter(is_deleted=False, benefit_plan_id=benefit_plan_id, status="ACTIVE").distinct()
            beneficiary_already_active = active_beneficiaries.filter(id=id).exists() if id else False
            return active_beneficiaries.count() == max_active_beneficiaries and not beneficiary_already_active
        return False

    @register_service_signal('beneficiary_service.create')
    def create(self, obj_data):
        try:
            status = obj_data.get("status", None)
            benefit_plan_id = obj_data.get("benefit_plan_id", None)

            if self.would_exceed_max_active_beneficiaries(benefit_plan_id, status):
                raise ValueError(f"Error creating beneficiary with active status. Benefit plan is already at max active beneficiaries")
            return super().create(obj_data)
        except Exception as exc:
            return output_exception(model_name=self.OBJECT_TYPE.__name__, method="update", exception=exc)
    
    @register_service_signal('beneficiary_service.update')
    def update(self, obj_data):
        try:
            status = obj_data.get("status", None)
            benefit_plan_id = obj_data.get("benefit_plan_id", None)
            id = obj_data.get('id', None)

            if self.would_exceed_max_active_beneficiaries(benefit_plan_id, status, id):
                raise ValueError(f"Error changing beneficiary to active status. Benefit plan is already at max active beneficiaries")
            return super().update(obj_data)
        except Exception as exc:
            return output_exception(model_name=self.OBJECT_TYPE.__name__, method="update", exception=exc)

    @register_service_signal('beneficiary_service.delete')
    def delete(self, obj_data):
        return super().delete(obj_data)

    def _business_data_serializer(self, data):
        def serialize(key, value):
            if key == 'id':
                beneficiary = Beneficiary.objects.get(id=value)
                individual = beneficiary.individual
                return f'{individual.first_name} {individual.last_name}'
            elif key == 'benefit_plan_id':
                benefit_plan = BenefitPlan.objects.get(id=value)
                return benefit_plan.__str__()
            elif key == 'individual_id':
                individual = Individual.objects.get(id=value)
                return f'{individual.first_name} {individual.last_name}'
            else:
                return value

        serialized_data = crud_business_data_builder(data, serialize)
        return serialized_data


class GroupBeneficiaryService(BaseService, CheckerLogicServiceMixin):
    OBJECT_TYPE = GroupBeneficiary

    def __init__(self, user, validation_class=GroupBeneficiaryValidation):
        super().__init__(user, validation_class)

    def would_exceed_max_active_beneficiaries(self, benefit_plan_id, status, id=None):
        benefit_plan = BenefitPlan.objects.get(id=benefit_plan_id)
        if benefit_plan and status == "ACTIVE":
            max_active_beneficiaries = benefit_plan.max_beneficiaries
            active_beneficiaries = GroupBeneficiary.objects.filter(is_deleted=False, benefit_plan_id=benefit_plan_id, status="ACTIVE").distinct()
            beneficiary_already_active = active_beneficiaries.filter(id=id).exists() if id else False
            return active_beneficiaries.count() == max_active_beneficiaries and not beneficiary_already_active
        return False

    @register_service_signal('group_beneficiary_service.create')
    def create(self, obj_data):
        try:
            status = obj_data.get("status", None)
            benefit_plan_id = obj_data.get("benefit_plan_id", None)

            if self.would_exceed_max_active_beneficiaries(benefit_plan_id, status):
                raise ValueError(f"Error creating beneficiary with active status. Benefit plan is already at max active beneficiaries")
            return super().create(obj_data)
        except Exception as exc:
            return output_exception(model_name=self.OBJECT_TYPE.__name__, method="update", exception=exc)

    @register_service_signal('group_beneficiary_service.update')
    def update(self, obj_data):
        try:
            status = obj_data.get("status", None)
            benefit_plan_id = obj_data.get("benefit_plan_id", None)
            id = obj_data.get('id', None)

            if self.would_exceed_max_active_beneficiaries(benefit_plan_id, status, id):
                raise ValueError(f"Error changing beneficiary to active status. Benefit plan is already at max active beneficiaries")
            return super().update(obj_data)
        except Exception as exc:
            return output_exception(model_name=self.OBJECT_TYPE.__name__, method="update", exception=exc)

    @register_service_signal('group_beneficiary_service.delete')
    def delete(self, obj_data):
        return super().delete(obj_data)


class BeneficiaryImportService:
    import_loaders = {
        # .csv
        'text/csv': lambda f: pd.read_csv(f),
        # .xlsx
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': lambda f: pd.read_excel(f),
        # .xls
        'application/vnd.ms-excel': lambda f: pd.read_excel(f),
        # .ods
        'application/vnd.oasis.opendocument.spreadsheet': lambda f: pd.read_excel(f),
    }

    def __init__(self, user):
        super().__init__()
        self.user = user

    @register_service_signal('benefit_plan.import_beneficiaries')
    def import_beneficiaries(self,
                             import_file: InMemoryUploadedFile,
                             benefit_plan: BenefitPlan,
                             workflow: WorkflowHandler,
                             group_aggregation_column: str):
        upload = self._save_sources(import_file)
        self._create_benefit_plan_data_upload_records(benefit_plan, workflow, upload, group_aggregation_column)
        self._trigger_workflow(workflow, upload, benefit_plan)
        return {'success': True, 'data': {'upload_uuid': upload.uuid}}

    @transaction.atomic
    def _save_sources(self, import_file):
        # Method separated as workflow execution must be independent of the atomic transaction.
        upload = self._create_upload_entry(import_file.name)
        dataframe = self._load_import_file(import_file)
        self._validate_dataframe(dataframe)
        self._save_data_source(dataframe, upload)
        return upload

    @transaction.atomic
    def _create_benefit_plan_data_upload_records(self, benefit_plan, workflow, upload, group_aggregation_column):
        record = BenefitPlanDataUploadRecords(
            data_upload=upload,
            benefit_plan=benefit_plan,
            workflow=workflow.name,
            json_ext={"group_aggregation_column": group_aggregation_column}
        )
        record.save(user=self.user)

    def validate_import_beneficiaries(self, upload_id: uuid, individual_sources, benefit_plan: BenefitPlan):
        dataframe = self._load_dataframe(individual_sources)
        validated_dataframe, invalid_items = self._validate_possible_beneficiaries(
            dataframe,
            benefit_plan,
            upload_id
        )
        return {'success': True, 'data': validated_dataframe, 'summary_invalid_items': invalid_items}

    def create_task_with_importing_valid_items(self, upload_id: uuid, benefit_plan: BenefitPlan):
        if SocialProtectionConfig.enable_maker_checker_for_beneficiary_upload:
            BeneficiaryTaskCreatorService(self.user) \
                .create_task_with_importing_valid_items(upload_id, benefit_plan)
        else:
            record = BenefitPlanDataUploadRecords.objects.get(
                data_upload_id=upload_id,
                is_deleted=False
            )
            from social_protection.signals.on_validation_import_valid_items import IndividualItemsImportTaskCompletionEvent
            IndividualItemsImportTaskCompletionEvent(
                SocialProtectionConfig.validation_import_valid_items_workflow,
                record,
                record.data_upload.id,
                record.benefit_plan,
                self.user
            ).run_workflow()

    def create_task_with_update_valid_items(self, upload_id: uuid, benefit_plan: BenefitPlan):
        if SocialProtectionConfig.enable_maker_checker_for_beneficiary_update:
            BeneficiaryTaskCreatorService(self.user)\
                .create_task_with_update_valid_items(upload_id, benefit_plan)

        # Resolve automatically if maker-checker not enabled
        if not SocialProtectionConfig.enable_maker_checker_for_beneficiary_update:
            from social_protection.signals.on_validation_import_valid_items import IndividualItemsUploadTaskCompletionEvent
            record = BenefitPlanDataUploadRecords.objects.get(
                data_upload_id=upload_id,
                benefit_plan=benefit_plan,
                is_deleted=False
            )
            IndividualItemsUploadTaskCompletionEvent(
                SocialProtectionConfig.validation_upload_valid_items_workflow,
                record,
                record.data_upload.id,
                record.benefit_plan,
                self.user
            ).run_workflow()

    def synchronize_data_for_reporting(self, upload_id: uuid, benefit_plan: BenefitPlan):
        self._synchronize_individual(upload_id)
        self._synchronize_beneficiary(benefit_plan, upload_id)

    def _validate_possible_beneficiaries(self, dataframe: DataFrame, benefit_plan: BenefitPlan, upload_id: uuid):

        if isinstance(benefit_plan.beneficiary_data_schema, str):
            schema_dict = json.loads(benefit_plan.beneficiary_data_schema)
        else:
            schema_dict = benefit_plan.beneficiary_data_schema
        properties = schema_dict.get("properties", {})

        calculation_uuid = SocialProtectionConfig.validation_calculation_uuid
        calculation = get_calculation_object(calculation_uuid)

        unique_fields = [field for field, props in properties.items() if "uniqueness" in props]
        unique_validations = {}
        if unique_fields:
            unique_validations = {
                field: dataframe[field].duplicated(keep=False)
                for field in unique_fields
            }

        # TODO: Use ProcessPoolExecutor after resolving django dependency loading issue
        validated_dataframe = BeneficiaryImportService.process_chunk(
            dataframe,
            properties,
            unique_validations,
            calculation,
            calculation_uuid,
        )

        self.save_validation_error_in_data_source_bulk(validated_dataframe)
        invalid_items = fetch_summary_of_broken_items(upload_id)
        return validated_dataframe, invalid_items

    @staticmethod
    def process_chunk(chunk, properties, unique_validations, calculation, calculation_uuid):
        validated_dataframe = []
        for _, row in chunk.iterrows():
            field_validation = {'row': row.to_dict(), 'validations': {}}
            for field, field_properties in properties.items():

                # Validation Calculation
                if "validationCalculation" in field_properties and field in row:
                    validation_name = field_properties["validationCalculation"]["name"]
                    field_validation['validations'][field] = calculation.calculate_if_active_for_object(
                        validation_name,
                        calculation_uuid,
                        field_name=field,
                        field_value=row[field]
                    )

                # Uniqueness Check
                if "uniqueness" in field_properties and field in row:
                    field_validation['validations'][f'{field}_uniqueness'] = {
                        'success': not unique_validations[field].loc[row.name] 
                    }
                    

            validated_dataframe.append(field_validation)

        return validated_dataframe

    def _handle_uniqueness(self, row, field, field_properties, benefit_plan, dataframe):
        unique_class_validation = SocialProtectionConfig.unique_class_validation
        calculation_uuid = SocialProtectionConfig.validation_calculation_uuid
        calculation = get_calculation_object(calculation_uuid)
        result_row = calculation.calculate_if_active_for_object(
            unique_class_validation,
            calculation_uuid,
            field_name=field,
            field_value=row[field],
            benefit_plan=benefit_plan.id,
            incoming_data=dataframe
        )
        return result_row

    def _handle_validation_calculation(self, row, field, field_properties):
        validation_calculation = field_properties.get("validationCalculation", {}).get("name")
        if not validation_calculation:
            raise ValueError("Missing validation name")
        calculation_uuid = SocialProtectionConfig.validation_calculation_uuid
        calculation = get_calculation_object(calculation_uuid)
        result_row = calculation.calculate_if_active_for_object(
            validation_calculation,
            calculation_uuid,
            field_name=field,
            field_value=row[field]
        )
        return result_row

    def _create_upload_entry(self, filename):
        upload = IndividualDataSourceUpload(source_name=filename, source_type='beneficiary import')
        upload.save(username=self.user.login_name)
        return upload

    def _validate_dataframe(self, dataframe: pd.DataFrame):
        if dataframe is None:
            raise ValueError("Unknown error while loading import file")
        if dataframe.empty:
            raise ValueError("Import file is empty")

    def _load_import_file(self, import_file) -> pd.DataFrame:
        if import_file.content_type not in self.import_loaders:
            raise ValueError("Unsupported content type: {}".format(import_file.content_type))

        return self.import_loaders[import_file.content_type](import_file)

    def _save_data_source(self, dataframe: pd.DataFrame, upload: IndividualDataSourceUpload):
        data_source_objects = []

        for _, row in dataframe.iterrows():
            ds = IndividualDataSource(
                upload=upload,
                json_ext=json.loads(row.to_json()),
                validations={},
                user_created=self.user,
                user_updated=self.user,
                uuid=uuid.uuid4()
            )
            data_source_objects.append(ds)

        IndividualDataSource.objects.bulk_create(data_source_objects)

    def _save_row(self, row, upload):
        ds = IndividualDataSource(upload=upload, json_ext=json.loads(row.to_json()), validations={})
        ds.save(username=self.user.login_name)

    def _load_dataframe(self, individual_sources) -> pd.DataFrame:
        return load_dataframe(individual_sources)

    def _trigger_workflow(self,
                          workflow: WorkflowHandler,
                          upload: IndividualDataSourceUpload,
                          benefit_plan: BenefitPlan):
        try:
            # Before the run in order to avoid racing conditions
            upload.status = IndividualDataSourceUpload.Status.TRIGGERED
            upload.save(username=self.user.login_name)

            result = workflow.run({
                # Core user UUID required
                'user_uuid': str(User.objects.get(username=self.user.login_name).id),
                'benefit_plan_uuid': str(benefit_plan.uuid),
                'upload_uuid': str(upload.uuid),
            })

            # Conditions are safety measure for workflows. Usually handles like PythonHandler or LightningHandler
            #  should follow this pattern but return type is not determined in workflow.run abstract.
            if result and isinstance(result, dict) and result.get('success') is False:
                raise ValueError(result.get('message', 'Unexpected error during the workflow execution'))
        except ValueError as e:
            upload.status = IndividualDataSourceUpload.Status.FAIL
            upload.error = {'workflow': str(e)}
            upload.save(username=self.user.login_name)
            return upload

    def save_validation_error_in_data_source_bulk(self, validated_dataframe):
        data_sources_to_update = []

        for field_validation in validated_dataframe:
            row = field_validation['row']
            error_fields = []

            for key, value in field_validation['validations'].items():
                if not value.get('success', False):
                    error_fields.append({
                        "field_name": value.get('field_name'),
                        "note": value.get('note')
                    })

            data_sources_to_update.append(
                IndividualDataSource(
                    id=row['id'],
                    validations={'validation_errors': error_fields}
                )
            )

        if data_sources_to_update:
            IndividualDataSource.objects.bulk_update(data_sources_to_update, ['validations'])

    def _synchronize_individual(self, upload_id):
        individuals_to_update = Individual.objects.filter(
            individualdatasource__upload=upload_id
        )
        for individual in individuals_to_update:
            synch_status = {
                'report_synch': 'true',
                'version': individual.version + 1,
            }
            if individual.json_ext:
                individual.json_ext.update(synch_status)
            else:
                individual.json_ext = synch_status
            individual.save(user=self.user)

    def _synchronize_beneficiary(self, benefit_plan, upload_id):
        unique_uuids = list((
            Beneficiary.objects
                .filter(benefit_plan=benefit_plan, individual__individualdatasource__upload_id=upload_id)
                .values_list('id', flat=True)
                .distinct()
        ))
        beneficiaries = Beneficiary.objects.filter(
            id__in=unique_uuids
        )
        for beneficiary in beneficiaries:
            synch_status = {
                'report_synch': 'true',
                'version': beneficiary.version + 1,
            }
            if beneficiary.json_ext:
                beneficiary.json_ext.update(synch_status)
            else:
                beneficiary.json_ext = synch_status
            beneficiary.save(user=self.user)


class BeneficiaryTaskCreatorService:

    def __init__(self, user):
        self.user = user

    def create_task_with_importing_valid_items(self, upload_id: uuid, benefit_plan: BenefitPlan):
        self._create_task(benefit_plan, upload_id, SocialProtectionConfig.validation_import_valid_items)

    def create_task_with_update_valid_items(self, upload_id: uuid, benefit_plan: BenefitPlan):
        self._create_task(benefit_plan, upload_id, SocialProtectionConfig.validation_upload_valid_items)

    @register_service_signal('socialProtection.update_task')
    @transaction.atomic()
    def _create_task(self, benefit_plan, upload_id, business_event):
        from social_protection.apps import SocialProtectionConfig
        from tasks_management.services import TaskService
        from tasks_management.apps import TasksManagementConfig
        from tasks_management.models import Task
        upload_record = BenefitPlanDataUploadRecords.objects.get(
            data_upload_id=upload_id,
            benefit_plan=benefit_plan,
            is_deleted=False
        )
        json_ext = {
            'benefit_plan_code': benefit_plan.code,
            'source_name': upload_record.data_upload.source_name,
            'workflow': upload_record.workflow,
            'percentage_of_invalid_items': self.__calculate_percentage_of_invalid_items(upload_id),
            'data_upload_id': str(upload_id)
        }
        TaskService(self.user).create({
            'source': 'import_valid_items',
            'entity': upload_record,
            'status': Task.Status.RECEIVED,
            'executor_action_event': TasksManagementConfig.default_executor_event,
            'business_event': business_event,
            'json_ext': json_ext
        })

        data_upload = upload_record.data_upload
        data_upload.status = IndividualDataSourceUpload.Status.WAITING_FOR_VERIFICATION
        data_upload.save(user=self.user)

    def __calculate_percentage_of_invalid_items(self, upload_id):
        number_of_valid_items = len(fetch_summary_of_valid_items(upload_id))
        number_of_invalid_items = len(fetch_summary_of_broken_items(upload_id))
        total_items = number_of_invalid_items + number_of_valid_items

        if total_items == 0:
            percentage_of_invalid_items = 0
        else:
            percentage_of_invalid_items = (number_of_invalid_items / total_items) * 100

        percentage_of_invalid_items = round(percentage_of_invalid_items, 2)
        return percentage_of_invalid_items


class GroupBeneficiaryImportService(BeneficiaryImportService):
    pass
    # TODO: create workflow upload/update groups and use it here


class LocationBenefitPlanPaymentPointService(BaseService):
    OBJECT_TYPE = LocationBenefitPlanPaymentPoint

    def __init__(self, user):
        super().__init__(user)

    @register_service_signal('location_benefit_plan_payment_point_service.create')
    def create(self, obj_data):
        return super().create(obj_data)

    @register_service_signal('location_benefit_plan_payment_point_service.update')
    def update(self, obj_data):
        return super().update(obj_data)

    @register_service_signal('location_benefit_plan_payment_point_service.delete')
    def delete(self, obj_data):
        return super().delete(obj_data)


class BeneficiaryExcelExportService:
    """Service for exporting beneficiaries to Excel with specific format including photo URLs"""
    
    def __init__(self, user, base_url='https://mis.merankabandi2.bi'):
        self.user = user
        self.base_url = base_url.rstrip('/')
    
    def export_group_beneficiaries_to_excel(self, group_beneficiaries_queryset, filename='beneficiaries.xlsx'):
        """Export group beneficiaries to Excel with the specified format"""
        import openpyxl
        from openpyxl import Workbook
        from django.http import HttpResponse
        from io import BytesIO
        from individual.models import GroupIndividual
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Beneficiaries"

        # Define headers
        headers = [
            'nom', 'prenom', 'province', 'commune', 'colline', 
            'cni', 'naissance_date', 'genre', 'socialid', 'pere', 'mere',
            'photo', 'cni_recto', 'cni_verso'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        row_num = 2
        
        # Iterate through group beneficiaries
        for group_beneficiary in group_beneficiaries_queryset.select_related(
            'group', 
            'group__location', 
            'group__location__parent',
            'group__location__parent__parent'
        ).prefetch_related('group__groupindividuals__individual'):
            
            # Get location hierarchy from group
            colline = group_beneficiary.group.location if group_beneficiary.group.location else None
            commune = colline.parent if colline and colline.parent else None
            province = commune.parent if commune and commune.parent else None
            
            # Get only the primary recipient from the group
            primary_recipient = GroupIndividual.objects.filter(
                group=group_beneficiary.group,
                recipient_type='PRIMARY',
                is_deleted=False
            ).select_related('individual').first()
            
            # Skip if no primary recipient found
            if not primary_recipient:
                continue
                
            # Export only the primary recipient
            group_individual = primary_recipient
            individual = group_individual.individual
            
            # Get gender from json_ext
            genre = individual.json_ext.get('sexe', '') if individual.json_ext else ''
            
            # Get CNI from json_ext
            cni = individual.json_ext.get('ci', '') if individual.json_ext else ''
            
            # Format date
            naissance_date = individual.dob.strftime('%Y-%m-%d') if individual.dob else ''
            socialid = group_beneficiary.group.code if group_beneficiary.group else ''
            pere = individual.json_ext.get('pere', '') if individual.json_ext else ''
            mere = individual.json_ext.get('mere', '') if individual.json_ext else ''

            # Generate photo URLs
            individual_uuid = str(individual.id)
            photo_url = f"{self.base_url}/api/merankabandi/beneficiary-photo/photo/{individual_uuid}/"
            cni_recto_url = f"{self.base_url}/api/merankabandi/beneficiary-photo/photo_ci1/{individual_uuid}/"
            cni_verso_url = f"{self.base_url}/api/merankabandi/beneficiary-photo/photo_ci2/{individual_uuid}/"
            
            # Write row data
            row_data = [
                individual.last_name,
                individual.first_name,
                province.name if province else '',
                commune.name if commune else '',
                colline.name if colline else '',
                cni,
                naissance_date,
                genre,
                socialid,
                pere,
                mere,
                photo_url,
                cni_recto_url,
                cni_verso_url
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)
            
            row_num += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file

class ProjectService(BaseService):
    OBJECT_TYPE = Project

    def __init__(self, user, validation_class=ProjectValidation):
        super().__init__(user, validation_class)

    @register_service_signal("project_service.create")
    def create(self, obj_data):
        return super().create(obj_data)

    @register_service_signal("project_service.update")
    def update(self, obj_data):
        return super().update(obj_data)

    @register_service_signal("project_service.delete")
    def delete(self, obj_data):
        return super().delete(obj_data)

    @register_service_signal('project_service.undo_delete')
    def undo_delete(self, obj_data):
        try:
            with transaction.atomic():
                self.validation_class.validate_undo_delete(obj_data)
                obj_ = self.OBJECT_TYPE.objects.filter(id=obj_data['id']).first()
                obj_.is_deleted = False
                obj_.save(user=self.user.user)
                return {
                    "success": True,
                    "message": "Ok",
                    "detail": "Undo Delete",
                }
        except Exception as exc:
            return output_exception(
                model_name=self.OBJECT_TYPE.__name__, method="undo_delete", exception=exc
            )


